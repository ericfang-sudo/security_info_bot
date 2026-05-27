from __future__ import annotations

import base64
import re
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from src.utils.logging import log

IP_PATTERN = re.compile(
    r"\b(?:(?:25[0-5]|2[0-4]\d|[01]?\d\d?)\.){3}(?:25[0-5]|2[0-4]\d|[01]?\d\d?)\b"
)
HASH_PATTERN = re.compile(r"^[0-9a-fA-F]{32,64}$")
DOMAIN_PATTERN = re.compile(r"^[a-zA-Z0-9._-]+\.[a-zA-Z]{2,}$")

_SKIP_SHEETS = {"MITRE ATT&CK"}


def parse_xlsx_iocs(
    xlsx_bytes: bytes,
) -> tuple[list[str], list[str], list[str]]:
    """Parse an xlsx file and return (ips, hashes, domains) — each deduped, order-preserved."""
    seen_ips: dict[str, None] = {}
    seen_hashes: dict[str, None] = {}
    seen_domains: dict[str, None] = {}

    try:
        wb = load_workbook(
            filename=__import__("io").BytesIO(xlsx_bytes), read_only=True, data_only=True
        )
    except Exception as e:
        log.error("Failed to open xlsx bytes: %s", e)
        return [], [], []

    for ws in wb.worksheets:
        if ws.title in _SKIP_SHEETS:
            continue
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue
                val = str(cell).strip()
                if not val or val.lower() == "none":
                    continue
                # IP — may appear anywhere in cell (e.g. "192.168.1.1:8080")
                ip_hits = IP_PATTERN.findall(val)
                if ip_hits:
                    for ip in ip_hits:
                        seen_ips.setdefault(ip, None)
                    continue
                # Hash — cell must be exactly a hex string of right length
                if HASH_PATTERN.match(val):
                    seen_hashes.setdefault(val.lower(), None)
                    continue
                # Domain — whole cell, no spaces
                if " " not in val and DOMAIN_PATTERN.match(val):
                    seen_domains.setdefault(val.lower(), None)

    wb.close()
    return list(seen_ips), list(seen_hashes), list(seen_domains)


def extract_iocs_from_info_file(
    info_file: list[dict],
) -> tuple[list[str], list[str], list[str]]:
    """Extract IoCs from infoFile entries (TWCERT detail API).

    Each entry may have a ``file`` field containing a data URI:
    ``data:<mime>;base64,<b64>``.  Only .xlsx/.xls attachments are processed.
    Results from multiple attachments are merged and deduped.
    """
    all_ips: dict[str, None] = {}
    all_hashes: dict[str, None] = {}
    all_domains: dict[str, None] = {}

    for entry in info_file:
        fn = (entry.get("fileName") or "").lower()
        if not (fn.endswith(".xlsx") or fn.endswith(".xls")):
            continue
        data_uri = entry.get("file") or ""
        if not data_uri.startswith("data:"):
            continue
        try:
            b64_part = data_uri.split(",", 1)[1]
            xlsx_bytes = base64.b64decode(b64_part)
        except Exception as e:
            log.warning("Failed to decode base64 attachment %s: %s", entry.get("fileName"), e)
            continue

        ips, hashes, domains = parse_xlsx_iocs(xlsx_bytes)
        for v in ips:
            all_ips.setdefault(v, None)
        for v in hashes:
            all_hashes.setdefault(v, None)
        for v in domains:
            all_domains.setdefault(v, None)

    return list(all_ips), list(all_hashes), list(all_domains)


def write_ioc_txt(
    intel_id: str,
    ips: list[str],
    hashes: list[str],
    domains: list[str],
) -> Path | None:
    """Write IoCs to a temp .txt file with [IPs] / [Hashes] / [Domains] sections.

    Only non-empty sections are written. Returns None if all three lists are empty.
    """
    sections: list[str] = []
    if ips:
        sections.append("[IPs]\n" + "\n".join(sorted(set(ips))))
    if hashes:
        sections.append("[Hashes]\n" + "\n".join(sorted(set(hashes))))
    if domains:
        sections.append("[Domains]\n" + "\n".join(sorted(set(domains))))

    if not sections:
        return None

    safe_id = re.sub(r"[^\w\-]", "_", intel_id)
    output_path = Path(tempfile.gettempdir()) / f"ioc_{safe_id}.txt"
    output_path.write_text("\n\n".join(sections) + "\n", encoding="utf-8")
    log.info(
        "Wrote IoC txt for %s: %d IPs, %d hashes, %d domains → %s",
        intel_id,
        len(ips),
        len(hashes),
        len(domains),
        output_path,
    )
    return output_path
